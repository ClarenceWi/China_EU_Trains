import re
import openpyxl
import pdfplumber
from collections import defaultdict

def log(message):
    """打印日志信息"""
    print(f"[处理日志] {message}")

def extract_declarations_from_pdf(pdf_path):
    """直接从PDF文件提取报关单信息"""
    try:
        declarations = []
        
        with pdfplumber.open(pdf_path) as pdf:
            log(f"打开PDF文件: {pdf_path}，共 {len(pdf.pages)} 页")
            
            # 遍历所有页面
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                if not text:
                    continue
                    
                # 查找预录入编号
                dec_matches = re.findall(r'预录入编号[:：]\s*(\d{18})', text, re.IGNORECASE)
                if not dec_matches:
                    continue
                    
                # 每个预录入编号对应一个报关单
                for dec_num in dec_matches:
                    # 提取箱号
                    container_patterns = [
                        re.compile(r'箱号[:：]\s*([A-Z]{4}\d{7})', re.IGNORECASE),
                        re.compile(r'集装箱号[:：]\s*([A-Z]{4}\d{7})', re.IGNORECASE),
                        re.compile(r'[A-Z]{4}\d{7}'),  # 通用模式
                    ]
                    
                    containers = []
                    for pattern in container_patterns:
                        containers.extend(pattern.findall(text))
                    
                    # 去重箱号
                    containers = list(dict.fromkeys(containers))
                    
                    declarations.append({
                        'declaration_number': dec_num,
                        'containers': containers,
                        'page': page_num
                    })
                    
                    log(f"第 {page_num} 页: 提取报关单 {dec_num}，包含 {len(containers)} 个箱号")
        
        return declarations
        
    except Exception as e:
        log(f"PDF提取错误: {str(e)}")
        return []

def process_excel_directly(excel_path, pdf_path, container_col='F'):
    """直接处理Excel和PDF文件，不使用中间提取文件"""
    try:
        # 步骤1: 直接从PDF提取报关单信息
        log("\n===== 步骤1: 从PDF提取报关单信息 =====")
        declarations = extract_declarations_from_pdf(pdf_path)
        
        if not declarations:
            log("错误：未能从PDF中提取到报关单信息")
            return False
            
        log(f"成功提取 {len(declarations)} 个报关单信息")
        
        # 步骤2: 构建映射关系
        log("\n===== 步骤2: 构建箱号-报关单映射 =====")
        container_map = defaultdict(list)  # 箱号: [报关单号列表]
        declaration_map = {}               # 报关单号: [箱号列表]
        
        for dec in declarations:
            dec_num = dec['declaration_number']
            containers = dec['containers']
            declaration_map[dec_num] = containers
            
            for container in containers:
                container_map[container].append(dec_num)
        
        # 去重报关单号列表
        for container, decs in container_map.items():
            container_map[container] = list(dict.fromkeys(decs))
            
        log(f"构建了 {len(container_map)} 个箱号的映射关系")
        
        # 步骤3: 直接处理Excel文件
        log("\n===== 步骤3: 处理Excel文件 =====")
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        log(f"打开Excel工作表: {ws.title}，使用 {container_col} 列作为箱号列")
        
        results = []
        row = 4  # 从G4开始
        updated_count = 0
        no_match_count = 0
        rule1_count = 0
        rule2_count = 0
        
        while True:
            # 获取当前行的箱号
            container_cell = f"{container_col}{row}"
            try:
                container_value = ws[container_cell].value
            except:
                container_value = None
                
            if container_value is None:
                break  # 遇到空值停止
                
            # 标准化箱号
            container_value = str(container_value).replace('-', '').replace(' ', '').upper()
            
            # 跳过表头
            if container_value in ['箱号', '集装箱号', '出口']:
                row += 1
                continue
                
            # 查找匹配的报关单号
            if container_value in container_map:
                dec_numbers = container_map[container_value]
                
                if len(dec_numbers) > 1:
                    # 规则2: 一个箱号对应多个报关单号，用\连接
                    dec_value = '\\'.join(dec_numbers)
                    rule_type = 2
                    rule2_count += 1
                else:
                    # 规则1: 一个箱号对应一个报关单号
                    dec_value = dec_numbers[0]
                    rule_type = 1
                    rule1_count += 1
                    
                # 写入报关单号到G列
                ws[f"G{row}"] = dec_value
                updated_count += 1
                results.append({
                    'row': row,
                    'container': container_value,
                    'declaration': dec_value,
                    'rule': rule_type
                })
                log(f"行 {row}: 箱号 {container_value} 应用规则{rule_type}，填入 {dec_value}")
            else:
                # 未找到匹配
                ws[f"G{row}"] = ""
                no_match_count += 1
                results.append({
                    'row': row,
                    'container': container_value,
                    'declaration': '未匹配',
                    'rule': 0
                })
                log(f"行 {row}: 箱号 {container_value} 未找到匹配的报关单号")
                
            row += 1
        
        # 保存Excel文件
        wb.save(excel_path)
        log(f"Excel文件保存完成，共处理 {row-4} 行，更新 {updated_count} 行，未匹配 {no_match_count} 行")
        log(f"规则应用统计: 规则1={rule1_count}, 规则2={rule2_count}")
        
        # 生成结果报告（仅用于记录，不用于数据处理）
        generate_process_report(results, declarations, "直接处理结果报告.md")
        
        return True
        
    except Exception as e:
        log(f"处理错误: {str(e)}")
        return False

def generate_process_report(results, declarations, report_path):
    """生成处理报告"""
    try:
        with open(report_path, "w", encoding="utf-8") as f:
            f.write("# 报关单直接处理结果报告\n\n")
            
            # 统计信息
            total = len(results)
            rule1 = sum(1 for r in results if r['rule'] == 1)
            rule2 = sum(1 for r in results if r['rule'] == 2)
            no_match = sum(1 for r in results if r['rule'] == 0)
            
            f.write("## 处理统计\n")
            f.write(f"| 总处理行数 | 规则1匹配 | 规则2匹配 | 未匹配 |\n")
            f.write(f"|------------|-----------|-----------|--------|\n")
            f.write(f"| {total} | {rule1} | {rule2} | {no_match} |\n\n")
            
            # 未匹配项
            if no_match > 0:
                f.write("## 未匹配箱号\n")
                for r in results:
                    if r['rule'] == 0:
                        f.write(f"- 行 {r['row']}: {r['container']}\n")
                f.write("\n")
                
            # 报关单摘要
            f.write("## 报关单摘要\n")
            for dec in declarations[:5]:  # 只显示前5个
                f.write(f"- {dec['declaration_number']}: {len(dec['containers'])}个箱号 (第{dec['page']}页)\n")
            if len(declarations) > 5:
                f.write(f"- ... 共 {len(declarations)} 个报关单\n")
                
    except Exception as e:
        log(f"生成报告错误: {str(e)}")

if __name__ == "__main__":
    log("===== 报关单信息录入工具 =====")
    
    # 直接使用用户上传的原始文件路径
    pdf_path = input("请输入PDF文件路径: ")    # 原始PDF文件
    excel_path = input("请输入Excel文件路径: ")  # 原始Excel文件
    
    # 执行直接处理流程
    success = process_excel_directly(excel_path, pdf_path, container_col='F')
    
    if success:
        log("\n===== 处理完成 =====")
        log(f"处理结果已直接写入Excel文件: {excel_path}")
    else:
        log("\n===== 处理失败 =====")